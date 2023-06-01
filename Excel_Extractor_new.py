# -*- coding: utf-8 -*-
"""
Created on Tue May 30 10:44:24 2023

@author: ritbhatt
"""


import tkinter as tk
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import pandas as pd
import openpyxl
import re
from PyQt5.QtWidgets import QApplication, QWidget, QCalendarWidget, QVBoxLayout
from PyQt5.QtCore import QDate

# Load all sheets from the first Excel file into a list of dataframes
dfs1 = []

# Load all sheets from the second Excel file into a list of dataframes
dfs2 = []

sheet_name_list=[]
sheet_name_list1=[]
sheet_name_list2=[]

variable_name=[]
variable_value=[]
variable_font=[]
font_format_size=(("Arial", 34,"bold"),("Helvetica", 24),("Arial", 12),("Arial", 38,"bold"),("Helvetica", 16))

variable_colour=[]
list_of_colour=["blue","red","green","brown"]
date_no_list=["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]
dt_list=["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]
month_last_day_list=["31","28","31","30","31","30","30","31","30","31","30","31"]
month_no_list=["01","02","03", "04","05","06","07","08","09","10","11","12"]
mo_list=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
month_characterized_list=["January","February","March","April","May","June","July","August","September","October","November","December"]
year_no_list=["22","23"]
year_list=["2022","2023"]

wb1 = openpyxl.load_workbook("file1.xlsx")
sheet_names1 = wb1.sheetnames
for i in sheet_names1:
    sheet_name_list.append(i)
    sheet_name_list1.append(i)
wb1.save('file1.xlsx')


wb2 = openpyxl.load_workbook("file2.xlsx")
sheet_names2 = wb2.sheetnames
for i in sheet_names2:
    sheet_name_list.append(i)
    sheet_name_list2.append(i)
wb2.save('file2.xlsx')


app_sheet_list_Final = ['Log4J v1 Applications', 'A - Architect', 'A - Admin', 'C - Corp', 'B-Billing', 'HR Systems', 'Maintenance', 'DSR', 'QueryTracker', 'Jars', 'Capgemini Migration - Java Sec ', 'Comments', 'Summary']
app_sheet = ['A - Architect', 'A - Admin', 'C - Corp', 'B-Billing', 'HR Systems', 'Maintenance']
smart_columns_need_list=['Health', 'Application', 'App Status', 'App Group', 'AO Manager', 'AO Assigned', 'File Total Count', 'Log4j?', 'Secrets?', 'Bamboo Plan is Java v1.8? REQUIRED (AO)', 'Bamboo?', 'BitBucket?', 'Maven?', 'Module (Not Confirmed)', 'Project Completion Date', 'CAP Migration (y/n)', 'Original Scope', 'Start Date for Major Blocker', 'Start Date for Waiting On AO']
log4j_columns_need_list=['application_name', 'Notes', 'type', 'Install Location', 'CAP Migration (y/n)', 'RETIRED', 'Bamboo Project Key', 'Web file name', 'Module', 'Application Group', 'Numara team assignment ', 'Manager', 'Primary Contact', 'Secondary Contact', 'Business Process Group', 'Definition', 'Application Instance']
wbs_architect_columns_need_list=['Application_name', 'Type', 'CG App Owner', 'Artifactory Jars', 'IsSecrets', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'Remarks', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
wbs_admin_columns_need_list=['Application_name', 'Type', 'Artifactory Jars', 'CG App Owner', 'IsSecrets', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'REMARKS', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
wbs_corp_columns_need_list=['Application_name', 'Type', 'Artifactory Jars', 'CG App Owner', 'IsSecrets', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'Remarks', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
wbs_billing_columns_need_list=['Application_name', 'Type', 'Bamboo project key', 'Artifactory Jars', 'CG App Owner', 'IsSecrets', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'Remarks', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
wbs_hr_columns_need_list= ['Application_name', 'Type', 'Bamboo project key', 'Artifactory Jars', 'CG App Owner', 'IsSecrets', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'Remarks', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
wbs_maintenance_columns_need_list= ['Application_name', 'Type', 'isRetired', 'Bamboo project key', 'CICD Owner', 'IsSecrets', 'Artifactory Jars', 'CG App Owner', 'Confluence Link', 'IsConfluenceVerified', 'File Count', 'Remarks', 'Build spec', 'Deploy spec', 'Spec Owner', 'Deployment status']
querytracker_columns_need_list=['Date', 'Query / Action Item', 'Raised By', 'Owner', 'Seveority', 'Status', 'Remarks']
jars_columns_need_list=['Jar Name', 'Application Name', 'CG AO', 'IsIntenal', 'Bamboo Key', 'Group ID', 'Artifact ID', 'Version', 'Location', 'Status', 'Remarks']
dsr_columns_need_list=['Date', 'App Name', 'CG App Owner', 'Work Type', 'Action Item', 'Action Owner', 'Planned Date', 'Status', 'Remarks']

All_apps=[]

column_names_jar=[]

column_names_smart=[]

Apps_smart=[]

Jars_sheet=[]


app_type = ['Architect', 'Admin', 'Corp', 'Billing', 'HR', 'Maintenance']
Apps=[]


search_row_no_list=[]


def common_things(home_val):
    wb_1 = load_workbook("file1.xlsx")
    sheets = wb_1.sheetnames
    
    


    for sheet_name in app_sheet:
        sheet= wb_1[sheet_name]
        for cell in sheet.iter_cols(min_row=2,
                               max_row=None,
                               min_col=1,
                               max_col=1):
            Apps.append([data.value for data in cell])

    for i in Apps:
        for j in i:
            All_apps.append(j)
            
    All_apps1=[value_11 for value_11 in All_apps if value_11 is not None]
    All_apps2 = All_apps1.sort()

    #all jars

    jar_sheet_no=app_sheet_list_Final[9]
    sheetjar = wb_1[jar_sheet_no]

    # Get the maximum column index
    max_col_jar = sheetjar.max_column
    # Loop through the columns and print their names
    for col_jar in range(1, max_col_jar+1):
        col_name_jar = sheetjar.cell(row=1, column=col_jar).value
        column_names_jar.append(col_name_jar)



    sheet_name_jar = app_sheet_list_Final[9]
    sheet_jar= wb_1[sheet_name_jar]
    for cell1 in sheet_jar.iter_cols(min_row=2,
                               max_row=None,
                               min_col=1,
                               max_col=1):
        Jars_sheet.append([data.value for data in cell1])

    All_jars_sheet=Jars_sheet[0]

    All_jars_sheet1 =[value_111 for value_111 in All_jars_sheet if value_111 is not None]
    All_jars_sheet2 = All_jars_sheet1.sort()
    
    def give_a_space():
        tspace0 = " "
        variable_name.append(tspace0)
        variable_value.append(tspace0)
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
    
    def check(appname,Apps_1,app_sheet1):
        count_1=0
        app_index=0
        for i in Apps_1:
            if appname in i:
                x=count_1
                app_index=i.index(appname)
                break
            count_1+=1
        
        if count_1==0:
            #print(appgroup)
            give_a_space()
            give_a_space()
            
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            architect(app_sheet,count_1,app_index)
        elif count_1==1:
            #print(appgroup)
            give_a_space()
            give_a_space()
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            admin(app_sheet,count_1,app_index)
        elif count_1==2:
            #print(appgroup)
            give_a_space()
            give_a_space()
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            corp(app_sheet,count_1,app_index)
        elif count_1==3:
            #print(appgroup)
            give_a_space()
            give_a_space()
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            billing(app_sheet,count_1,app_index)
        elif count_1==4:
            #print(appgroup)
            give_a_space()
            give_a_space()
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            hr(app_sheet,count_1,app_index)
        elif count_1==5:
            #print(appgroup)
            give_a_space()
            give_a_space()
            t1 = " "
            t2 = app_sheet1[count_1]
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details"
            text2="Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            maintenance(app_sheet,count_1,app_index)
        else:
            #print(appgroup)
            give_a_space()
            give_a_space()
            texts1="               APPLICATION DETAILS FOR "
            ts2= appname
            variable_name.append(texts1)
            variable_value.append(ts2)
            font_size=font_format_size[3]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            text1="         WBS Sheet Details      "
            t1 = " "
            t2 = "Application Group is not belongs to Architect, Admin, Billing, Corp, HR, Maintenance "
            text1="  WBS Sheet Details  "
            text2="  Application Group :"
            variable_name.append(text1)
            variable_value.append(t1)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            variable_name.append(text2)
            variable_value.append(t2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[1]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            return 0
            




    def architect(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2 #row no for the appname
        
        # foolow this thing to remove dependency from coloumn no
        
        max_col_number = sheet1.max_column

        architect_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            architect_columns_list.append(col_name)

        
        # Penske appowner
        item=wbs_architect_columns_need_list[0] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[1] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[2] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[3] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[4] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[5] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[6] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[7] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[8] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[9] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[10] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[11] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_architect_columns_need_list[12] 
        col=architect_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        




    def admin(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2 
        
        max_col_number = sheet1.max_column

        admin_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            admin_columns_list.append(col_name)

        
        # Penske appowner
        item=wbs_admin_columns_need_list[0] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)

        
        # Penske appowner
        item=wbs_admin_columns_need_list[1] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[3] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[2] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[4] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[5] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[6] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[7] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[8] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[9] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[10] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[11] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_admin_columns_need_list[12] 
        col=admin_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        



    def billing(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2 
        
        
        max_col_number = sheet1.max_column

        billing_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            billing_columns_list.append(col_name)

        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[0] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[1] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[4] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[3] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[5] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[6] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[7] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[8] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[9] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[10] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[11] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[12] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_billing_columns_need_list[13] 
        col=billing_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
    
    
    def hr(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2 
        
        
        max_col_number = sheet1.max_column

        hr_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            hr_columns_list.append(col_name)

        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[0] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[1] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[4] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[3] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[5] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[6] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[7] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[8] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[9] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[10] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[11] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[12] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_hr_columns_need_list[13] 
        col=hr_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
    def maintenance(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2 
        
        
        max_col_number = sheet1.max_column

        maintenance_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            maintenance_columns_list.append(col_name)

        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[0] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[1] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[7] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[6] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[5] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[8] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[9] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[10] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[11] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[12] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[13] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[14] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_maintenance_columns_need_list[15] 
        col=maintenance_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
    
    
    
    
        
    def corp(app_sheet,count_1,app_index):
        sheet_name1=app_sheet[count_1]
        sheet1= wb_1[sheet_name1]
        row = app_index+2
        
        max_col_number = sheet1.max_column

        corp_columns_list=[]

        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            corp_columns_list.append(col_name)
        
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[0] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Name :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[1] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Application Type :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[3] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="CG App Owner/Developer :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[2] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Artifactory Jars :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[4] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsSecrets :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[5] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Confluence Link :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[6] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="IsConfluenceVerified :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[7] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="File Count :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[8] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Remarks :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[9] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Build spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[10] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deploy spec :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[11] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Spec Owner :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Penske appowner
        item=wbs_corp_columns_need_list[12] 
        col=corp_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t11 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t11==None):
            t11="No Data Available"
            font_colour=list_of_colour[1]
        text11="Deployment status :"
        variable_name.append(text11)
        variable_value.append(t11)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        



    def querytracker(appname):
        z=sheets.index(app_sheet_list_Final[8])
        sheet1= wb_1[sheets[z]] # sheet1= wb[sheets[10]]
        query_list=[]
        query_indexes=[]
        for cell in sheet1.iter_cols(min_row=2,
                               max_row=None,
                               min_col=1,
                               max_col=1):
            query_list.append([data.value for data in cell])
        i=query_list[0]
        
        s=0
        d=i.count(appname)
        while(d>0):
            query_indexes.append(i.index(appname)+s)
            i.remove(appname)
            s+=1
            d=i.count(appname)
        
        
        
        max_col_number = sheet1.max_column

        querytracker_columns_list=[]

        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            querytracker_columns_list.append(col_name)

        
        
        give_a_space()
        give_a_space()
        give_a_space()
        
        tspace1 = "         QUERYTRACKER DETAILS        "
        tspaceval = "        "
        variable_name.append(tspace1)
        variable_value.append(tspaceval)
        font_size=font_format_size[0]
        font_colour=list_of_colour[0]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        give_a_space()
        give_a_space()
        
        total_query = len(query_indexes)
        text_query="Total no of queries in Querytracker :"
        variable_name.append(text_query)
        variable_value.append(total_query)
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        c=1
        for query_index in query_indexes:
            row=query_index+2
            
            give_a_space()
            give_a_space()
            
            tqueryno = "Query No : "
            variable_name.append(tqueryno)
            variable_value.append(c)
            font_size=font_format_size[1]
            variable_font.append(font_size)


            item=querytracker_columns_need_list[0] 
            col=querytracker_columns_list.index(item) + 1
            val1 = sheet1.cell(row=row, column=col)
            t1 = val1.value
            font_size=font_format_size[2]
            if(t1==None):
                t1="No Data Available"
            text1="Query Date :"
            variable_name.append(text1)
            variable_value.append(t1)
            variable_font.append(font_size)
            
            
            # Query Status 
            item=querytracker_columns_need_list[5] 
            col=querytracker_columns_list.index(item) + 1
            val2 = sheet1.cell(row=row, column=col)
            t2 = val2.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            count1=0
            if(t2==None):
                font_colour=list_of_colour[1]
                count1=1
            elif(t2=="On Hold" or t2=="In Progress" or t2=="Open"):
                font_colour=list_of_colour[1]
                count1=1
            else:
                font_colour=list_of_colour[2]
            text2="Query Status :"
            variable_name.append(text2)
            variable_value.append(t2)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            font_colour=list_of_colour[2]
            variable_colour.append(font_colour)
            variable_colour.append(font_colour)
            
            
            # Query
            item=querytracker_columns_need_list[1] 
            col=querytracker_columns_list.index(item) + 1
            val3 = sheet1.cell(row=row, column=col)
            t3 = val3.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t3==None):
                t3="No Data Available"
                font_colour=list_of_colour[1]
            text3="Query :"
            variable_name.append(text3)
            variable_value.append(t3)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Raised By
            item=querytracker_columns_need_list[2] 
            col=querytracker_columns_list.index(item) + 1
            val4 = sheet1.cell(row=row, column=col)
            t4 = val4.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t4==None):
                t4="No Data Available"
                font_colour=list_of_colour[1]
            text4="Raised By :"
            variable_name.append(text4)
            variable_value.append(t4)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Raised To
            item=querytracker_columns_need_list[3] 
            col=querytracker_columns_list.index(item) + 1
            val5 = sheet1.cell(row=row, column=col)
            t5 = val5.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t5==None):
                t5="No Data Available"
                font_colour=list_of_colour[1]
            text5="Raised To :"
            variable_name.append(text5)
            variable_value.append(t5)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Seveority 
            item=querytracker_columns_need_list[4] 
            col=querytracker_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                t6="No Data Available"
                font_colour=list_of_colour[1]
            text6="Seveority :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Remarks 
            item=querytracker_columns_need_list[6] 
            col=querytracker_columns_list.index(item) + 1
            val7 = sheet1.cell(row=row, column=col)
            t7 = val7.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t7==None):
                t7="No Data Available"
                if(count1==1):
                    font_colour=list_of_colour[1]
            text7="Remarks :"
            variable_name.append(text7)
            variable_value.append(t7)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            c+=1
        

    def dsr(appname):
        z=sheets.index(app_sheet_list_Final[7])
        sheet1= wb_1[sheets[z]] # sheet1= wb[sheets[9]]
        dsr_list=[]
        dsr_indexes=[]
        for cell in sheet1.iter_cols(min_row=2,
                               max_row=None,
                               min_col=2,
                               max_col=2):
            dsr_list.append([data.value for data in cell])
        i=dsr_list[0]
        
        s=0
        d=i.count(appname)
        while(d>0):
            dsr_indexes.append(i.index(appname)+s)
            i.remove(appname)
            s+=1
            d=i.count(appname)
        
        
        
        # foolow this thing to remove dependency from coloumn no
        
        max_col_number = sheet1.max_column

        dsr_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            dsr_columns_list.append(col_name)

        
        
        give_a_space()
        give_a_space()
        give_a_space()
        
        
        tspace1 = "         DSR DETAILS        "
        tspaceval = "        "
        variable_name.append(tspace1)
        variable_value.append(tspaceval)
        font_size=font_format_size[0]
        font_colour=list_of_colour[0]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        give_a_space()
        give_a_space()
        
        total_dsr = len(dsr_indexes)
        text_dsr ="Total no of DSR :"
        variable_name.append(text_dsr)
        variable_value.append(total_dsr)
        font_size=font_format_size[1]
        font_colour=list_of_colour[3]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        c=1
        
        for dsr_index in dsr_indexes:
            
            row=dsr_index+2
        
            give_a_space()
            give_a_space()

            item=dsr_columns_need_list[3] 
            col=dsr_columns_list.index(item) + 1
            val1 = sheet1.cell(row=row, column=col)
            t1 = val1.value
            font_size=font_format_size[1]
            if(t1==None):
                t1="No Data Available"
            text1="DSR Work :"
            variable_name.append(text1)
            variable_value.append(t1)
            variable_font.append(font_size)
            
            
            # DSR Status 
            item=dsr_columns_need_list[7] 
            col=dsr_columns_list.index(item) + 1
            val2 = sheet1.cell(row=row, column=col)
            t2 = val2.value
            font_size=font_format_size[2]
            font_colour1=list_of_colour[2]
            if(t2==None):
                t2="No Data Available"
                font_colour=list_of_colour[1]
            elif(t2=="On Hold" or t2=="In Progress" or t2=="Open" or t2=="NA"):
                font_colour=list_of_colour[1]
            else:
                font_colour=list_of_colour[2]
            text2="Status :"
            variable_name.append(text2)
            variable_value.append(t2)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            variable_colour.append(font_colour1)
            
            
            # Date
            item=dsr_columns_need_list[0] 
            col=dsr_columns_list.index(item) + 1
            val3 = sheet1.cell(row=row, column=col)
            t3 = val3.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t3==None):
                t3="No Data Available"
                font_colour=list_of_colour[1]
            text3="Date :"
            variable_name.append(text3)
            variable_value.append(t3)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Action Owner
            item=dsr_columns_need_list[5] 
            col=dsr_columns_list.index(item) + 1
            val4 = sheet1.cell(row=row, column=col)
            t4 = val4.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t4==None):
                t4="No Data Available"
                font_colour=list_of_colour[1]
            text4="Action Owner :"
            variable_name.append(text4)
            variable_value.append(t4)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Action Item 
            item=dsr_columns_need_list[4] 
            col=dsr_columns_list.index(item) + 1
            val5 = sheet1.cell(row=row, column=col)
            t5 = val5.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t5==None):
                t5="No Data Available"
                font_colour=list_of_colour[1]
            text5="Action Item :"
            variable_name.append(text5)
            variable_value.append(t5)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            # Remarks 
            item=dsr_columns_need_list[8] 
            col=dsr_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                t6="No Data Available"
            text6="Remarks :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            c+=1
        
          
        


        
    def log4jv1(appname):
        z=sheets.index(app_sheet_list_Final[0])
        sheet1= wb_1[sheets[z]] # sheet1= wb[sheets[0]]
        log4j_list=[]
        for cell in sheet1.iter_cols(min_row=2,
                               max_row=None,
                               min_col=1,
                               max_col=1):
            log4j_list.append([data.value for data in cell])
        i=log4j_list[0]
        
        
        give_a_space()
        give_a_space()
        give_a_space()
        
        tspace1 = "         LOG4J DETAILS        "
        tspaceval = "        "
        variable_name.append(tspace1)
        variable_value.append(tspaceval)
        font_size=font_format_size[0]
        font_colour=list_of_colour[0]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        give_a_space()
        give_a_space()
        
        e=0
        if appname in i:
                app_index1=i.index(appname)
                e=1
        if(e==0):
            # print("No such app found in app list excel file")
            return 0

        
        row=app_index1+2
        
        max_col_number = sheet1.max_column

        log4j_columns_list=[]

        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            log4j_columns_list.append(col_name)
            
        
        # Penske App Owner
        item=log4j_columns_need_list[12] 
        col=log4j_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t1 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t1==None):
            t1="No Data Available"
            font_colour=list_of_colour[1]
        text1="App Owner(Penske) :"
        variable_name.append(text1)
        variable_value.append(t1)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Bamboo Project Key 
        item=log4j_columns_need_list[6] 
        col=log4j_columns_list.index(item) + 1
        val2 = sheet1.cell(row=row, column=col)
        t2 = val2.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t2==None):
            t2="No Data Available"
            font_colour=list_of_colour[1]
        text2="Bamboo Project Key :"
        variable_name.append(text2)
        variable_value.append(t2)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # cap migration
        item=log4j_columns_need_list[4] 
        col=log4j_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t1 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t1==None):
            t1="No Data Available"
            font_colour=list_of_colour[1]
        text1="Cap Migration :"
        variable_name.append(text1)
        variable_value.append(t1)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Retired  
        item=log4j_columns_need_list[5] 
        col=log4j_columns_list.index(item) + 1
        val2 = sheet1.cell(row=row, column=col)
        t2 = val2.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t2==None):
            t2="No Data Available"
            font_colour=list_of_colour[1]
        text2="Retired :"
        variable_name.append(text2)
        variable_value.append(t2)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Application Group(Excel)
        item=log4j_columns_need_list[9] 
        col=log4j_columns_list.index(item) + 1
        val1 = sheet1.cell(row=row, column=col)
        t1 = val1.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t1==None):
            t1="No Data Available"
            font_colour=list_of_colour[1]
        text1="Application Group(Excel) :"
        variable_name.append(text1)
        variable_value.append(t1)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        # Defination  
        item=log4j_columns_need_list[15] 
        col=log4j_columns_list.index(item) + 1
        val2 = sheet1.cell(row=row, column=col)
        t2 = val2.value
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        if(t2==None):
            t2="No Data Available"
            font_colour=list_of_colour[1]
        text2="Defination :"
        variable_name.append(text2)
        variable_value.append(t2)
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        

    def date_getter(date1):
        opdate=""
        pmon=date1[0:2]
        m1=month_no_list.index(pmon)
        m2=month_characterized_list[m1]
        pdat=date1[3:5]
        pyer=date1[6:8]
        y1=year_no_list.index(pyer)
        opdate+=str(m2)
        opdate+=" "
        opdate+=str(pdat)
        opdate+=", "
        opdate+=str(year_list[y1])
        return opdate 
    
    def jar(appname):
        z=sheets.index(app_sheet_list_Final[9])
        sheet1= wb_1[sheets[z]]
        jar_list=[]
        jar_indexes=[]
        for cell in sheet1.iter_cols(min_row=2,
                               max_row=None,
                               min_col=1,
                               max_col=2):
            jar_list.append([data.value for data in cell])
        i=jar_list[0] # JAR List
        j=jar_list[1] # application List for jar
        e=0
        if appname in i:
                jar_index1=i.index(appname)
                e=1
        if(e==0):
            J2=" "
            K2="No such jar found in jar list in excel file"
            variable_name.append(J2)
            variable_value.append(K2)
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            return 0
        row=jar_index1+2
        
        # foolow this thing to remove dependency from coloumn no
        
        s=0
        d=i.count(appname)
        while(d>0):
            jar_indexes.append(i.index(appname)+s)
            i.remove(appname)
            s+=1
            d=i.count(appname)
        
        max_col_number = sheet1.max_column

        jar_columns_list=[]

        # Loop through the columns and print their names
        for col_item in range(1, max_col_number+1):
            col_name = sheet1.cell(row=1, column=col_item).value
            jar_columns_list.append(col_name)

        
        
        give_a_space()
        give_a_space()
        give_a_space()
        
        tspace1 = "         JAR DETAILS        "
        tspaceval = "        "
        variable_name.append(tspace1)
        variable_value.append(tspaceval)
        font_size=font_format_size[0]
        font_colour=list_of_colour[0]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        give_a_space()
        give_a_space()
        
        total_jar = len(jar_indexes)
        text_jar="Total no of jars found in sheet :"
        variable_name.append(text_jar)
        variable_value.append(total_jar)
        font_size=font_format_size[2]
        font_colour=list_of_colour[2]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        ll1=len(jar_indexes)
        c=1
        for jar_index in jar_indexes:
            row=jar_index+2
            
            give_a_space()
            give_a_space()
            
            if(ll1>1):
                tjarno = "JAR No : "
                variable_name.append(tjarno)
                variable_value.append(c)
                font_size=font_format_size[0]
                font_colour=list_of_colour[3]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
            
            
            item=jars_columns_need_list[0] 
            col=jar_columns_list.index(item) + 1
            val1 = sheet1.cell(row=row, column=col)
            t1 = val1.value
            font_size=font_format_size[1]
            if(t1==None):
                t1="No Data Available"
            text1="Jar Name :"
            variable_name.append(text1)
            variable_value.append(t1)
            variable_font.append(font_size)
            
            
            item=jars_columns_need_list[9] 
            col=jar_columns_list.index(item) + 1
            val2 = sheet1.cell(row=row, column=col)
            t2 = val2.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t2==None):
                font_colour=list_of_colour[1]
            elif(t2=="On Hold" or t2=="In Progress" or t2=="Open"):
                font_colour=list_of_colour[1]
            else:
                font_colour=list_of_colour[2]
            text2="JAR Status :"
            variable_name.append(text2)
            variable_value.append(t2)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            font_colour=list_of_colour[2]
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[1] 
            col=jar_columns_list.index(item) + 1
            val3 = sheet1.cell(row=row, column=col)
            t3 = val3.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t3==None):
                t3="No Data Available"
                font_colour=list_of_colour[1]
            text3="JAR used in Application :"
            variable_name.append(text3)
            variable_value.append(t3)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[2] 
            col=jar_columns_list.index(item) + 1
            val4 = sheet1.cell(row=row, column=col)
            t4 = val4.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t4==None):
                t4="No Data Available"
                font_colour=list_of_colour[1]
            text4="CG Developer :"
            variable_name.append(text4)
            variable_value.append(t4)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[3] 
            col=jar_columns_list.index(item) + 1
            val5 = sheet1.cell(row=row, column=col)
            t5 = val5.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t5==None):
                t5="No Data Available"
                font_colour=list_of_colour[1]
            text5="Is JAR Intenal :"
            variable_name.append(text5)
            variable_value.append(t5)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[4] 
            col=jar_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            count=0
            if(t6==None):
                t6="No Data Available"
                count=1
            text6="Bamboo Key :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[5] 
            col=jar_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                if(count==1):
                    font_colour=list_of_colour[1]
                t6="No Data Available"
            text6="Group ID :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[6] 
            col=jar_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                if(count==1):
                    font_colour=list_of_colour[1]
                t6="No Data Available"
            text6="Artifact ID :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[7] 
            col=jar_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                if(count==1):
                    font_colour=list_of_colour[1]
                t6="No Data Available"
            text6="Version :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[8] 
            col=jar_columns_list.index(item) + 1
            val6 = sheet1.cell(row=row, column=col)
            t6 = val6.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t6==None):
                font_colour=list_of_colour[1]
                t6="No Data Available"
            text6="Location :"
            variable_name.append(text6)
            variable_value.append(t6)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=jars_columns_need_list[10] 
            col=jar_columns_list.index(item) + 1
            val7 = sheet1.cell(row=row, column=col)
            t7 = val7.value
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            if(t7==None):
                t7="No Data Available"
            text7="Remarks :"
            variable_name.append(text7)
            variable_value.append(t7)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            c+=1
    
    #smart apps

    wb_2 = load_workbook("file2.xlsx")
    sheets2 = wb_2.sheetnames
    smart_sheet_no=app_sheet_list_Final[10]
    sheet_2 = wb_2[smart_sheet_no]

    # Get the maximum column index
    max_col_smart = sheet_2.max_column
    # Loop through the columns and print their names
    for col_smart in range(1, max_col_smart+1):
        col_name_smart = sheet_2.cell(row=1, column=col_smart).value
        column_names_smart.append(col_name_smart)

    #smart apps

    applist_col_no=column_names_smart.index("Application")

    sheet_name_smart = app_sheet_list_Final[10]
    sheet_smart= wb_2[sheet_name_smart]
    row_count=1
    row_numbering_list=[]
    for cell1 in sheet_smart.iter_cols(min_row=2,
                               max_row=None,
                               min_col=applist_col_no+1,
                               max_col=applist_col_no+1):
        Apps_smart.append([data.value for data in cell1])
        

    All_apps_smart=Apps_smart[0]
    All_apps_smart1 =[value_111 for value_111 in All_apps_smart if value_111 is not None]
    All_apps_smart2 = All_apps_smart1.sort()
    
    
    for value_6 in All_apps_smart:
        row_name="Row "+str(row_count)
        row_numbering_list.append(row_name)
        row_count+=1
    

    #smart comments apps

    smart_sheet_comments_no=app_sheet_list_Final[11]
    sheet_3 = wb_2[smart_sheet_comments_no]
    
    column_number = 1  # Change to the appropriate column number
    column_letter = openpyxl.utils.get_column_letter(column_number)
    row_no_list=[]
    for cell in sheet_3[column_letter]:
        row_no_list.append(cell.value)
        
    new_list = [x for x in row_no_list if x is not None]
    
    column_number = 2  # Change to the appropriate column number
    column_letter = openpyxl.utils.get_column_letter(column_number)
    full_message_list=[]
    for cell in sheet_3[column_letter]:
        full_message_list.append(cell.value)
    
    column_number = 3  # Change to the appropriate column number
    column_letter = openpyxl.utils.get_column_letter(column_number)
    full_person_list=[]
    for cell in sheet_3[column_letter]:
        full_person_list.append(cell.value)
    
    date_column_number = 4  # Change to the appropriate column number
    date_column_letter = openpyxl.utils.get_column_letter(date_column_number)
    full_date_list=[]
    for date_cell in sheet_3[date_column_letter]:
        full_date_list.append(date_cell.value)
        
    
    date_list=[]
    el=0
    row_exact=row_no_list[0]
    for y in full_date_list:
        if(row_no_list[el] is not None):
            row_exact=row_no_list[el]
        if(y is not None):
            date_list.append(y[:8])
            search_row_no_list.append(row_exact)
        else:
            date_list.append(y)
            search_row_no_list.append(None)
        el+=1
            
    
            
    

    def smartsheet_comments(appnumber):
        
        e=0
        total_comments=0
        app_nam=row_numbering_list[appnumber]
        if app_nam in new_list:
            repeating_count=row_no_list.count(app_nam)
            e=1
            tspace0 = " "
            tspace1= str(repeating_count) + " comments present in Smartsheet "
            variable_name.append(tspace0)
            variable_value.append(tspace1)
            font_size=font_format_size[1]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
        if(e==0):
            textno1="No comments present in Smartsheet"
            tno1=" "
            variable_name.append(textno1)
            variable_value.append(tno1)
            font_size=font_format_size[1]
            font_colour=list_of_colour[1]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            return 0

        new_row_no_list=[]
        
        
        while repeating_count>=1:
            
            give_a_space()
            repeating_index_list = [app_number for app_number in range(len(row_no_list)) if row_no_list[app_number] == row_numbering_list[appnumber]]
            repeating_index=0
            limiting_index_list = []
            for every_value in repeating_index_list:
                every_value_1=every_value+1
                while(row_no_list[every_value_1] is None):
                    every_value_1+=1
                limit_value=every_value_1-every_value-2
                limiting_index_list.append(limit_value)
            repeating_count-=1
        give_a_space()
        lala=1
        for kk in range(len(repeating_index_list)):
            c=0
            p=limiting_index_list[kk]
            u=repeating_index_list[kk]
            give_a_space()
            give_a_space()
            text_name="Comment "
            text_value=str(lala)
            variable_name.append(text_name)
            variable_value.append(text_value)
            font_size=font_format_size[1]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            give_a_space()
            give_a_space()
            lala+=1
            for ll in range(u,u+p):
                if(c==0):
                    text_name=full_person_list[ll]+" asked on "+date_getter(full_date_list[ll])+" "
                else:
                    text_name=full_person_list[ll]+" replied on "+date_getter(full_date_list[ll])+" "
                
                
                
                text_value="\n"
                variable_name.append(text_name)
                variable_value.append(text_value)
                font_size=font_format_size[4]
                font_colour=list_of_colour[3]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                text_name=" "
                text_value=full_message_list[ll]
                variable_name.append(text_name)
                variable_value.append(text_value)
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                give_a_space()
                give_a_space()
                c+=1

        

    def smartsheet(appname):
        z=sheets2.index(app_sheet_list_Final[10])
        sheet1= wb_2[sheets2[z]] # sheet1= wb[sheets2[13]]
        smartapp_list=[]
        for cell in sheet1.iter_cols(min_row=2,
                               max_row=None,
                               min_col=applist_col_no+1,
                               max_col=applist_col_no+1):
            smartapp_list.append([data.value for data in cell])
        i=smartapp_list[0]
        
        
        give_a_space()
        give_a_space()
        give_a_space()
        
        tspace1 = "         SMARTSHEET DETAILS        "
        tspaceval = "        "
        variable_name.append(tspace1)
        variable_value.append(tspaceval)
        font_size=font_format_size[0]
        font_colour=list_of_colour[0]
        variable_font.append(font_size)
        variable_colour.append(font_colour)
        
        
        give_a_space()
        give_a_space()
        
        e=0
        repeat_app=0
        if appname in i:
            app_index1=All_apps_smart.index(appname)
            repeat_app=All_apps_smart.count(appname)
            e=1
        if(e==0):
            textno1="No such app found in app list in Smartsheet"
            tno1=" Please enter correct Smartsheet appname "
            variable_name.append(textno1)
            variable_value.append(tno1)
            font_size=font_format_size[1]
            font_colour=list_of_colour[1]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            return 0

        
        if repeat_app==1:
            row=app_index1+2
            # below things are not updated
            
            max_col_number = sheet1.max_column

            smart_columns_list=[]

            # Loop through the columns and print their names
            for col_item in range(1, max_col_number+1):
                col_name = sheet1.cell(row=1, column=col_item).value
                smart_columns_list.append(col_name)

            
            item=smart_columns_need_list[1] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Application Name :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[5] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="App Owner(Penske) :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[0] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="App Health :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[2] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="App Status :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[3] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="App Group :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[4] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="App Owner Manager(Penske) :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[6] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="File Total Count :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[7] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Log4j ?  "
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[8] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Secrets? "
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)

            
            item=smart_columns_need_list[10] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Bamboo? "
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[11] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="BitBucket? "
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[12] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Maven? "
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[13] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Module (Not Confirmed) :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[14] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Project Completion Date :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[15] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="CAP Migration (y/n) :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[16] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Original Scope :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[17] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Start Date for Major Blocker :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            item=smart_columns_need_list[18] 
            col=smart_columns_list.index(item) + 1
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            val1 = sheet1.cell(row=row, column=col)
            t11 = val1.value
            if(t11==None):
                t11="No Data Available"
                font_colour=list_of_colour[1]
            text11="Start Date for Waiting On AO :"
            variable_name.append(text11)
            variable_value.append(t11)
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            give_a_space()
            give_a_space()
            give_a_space()
            
            tspace1 = "         SMARTSHEET COMMENTS        "
            tspaceval = "        "
            variable_name.append(tspace1)
            variable_value.append(tspaceval)
            font_size=font_format_size[0]
            font_colour=list_of_colour[0]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            
            give_a_space()
            give_a_space()
            
            smartsheet_comments(app_index1)
            
        if repeat_app>1:
            
            # below things are not updated
            
            give_a_space()
            
            tspace0 = " "
            tspace1="App found in Smartsheet " + str(repeat_app) + " times"
            variable_name.append(tspace0)
            variable_value.append(tspace1)
            font_size=font_format_size[1]
            font_colour=list_of_colour[3]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            
            max_col_number = sheet1.max_column

            smart_columns_list=[]
            
            # repeated_app_list=[] fj
            
            repeated_app_list = [app_no for app_no in range(len(All_apps_smart)) if All_apps_smart[app_no] == appname]

            # Loop through the columns and print their names
            for col_item in range(1, max_col_number+1):
                col_name = sheet1.cell(row=1, column=col_item).value
                smart_columns_list.append(col_name)
            
            c=0
            for repeat_app_no in repeated_app_list:
                
                row = repeat_app_no + 2
                
                c+=1
                
                give_a_space()
                give_a_space()
                
                tspace0 = " "
                tspace1="App no found in Smartsheet: " + str(c)
                variable_name.append(tspace0)
                variable_value.append(tspace1)
                font_size=font_format_size[4]
                font_colour=list_of_colour[0]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                tspace0 = " "
                variable_name.append(tspace0)
                variable_value.append(tspace0)
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                item=smart_columns_need_list[1] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Application Name :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[5] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="App Owner(Penske) :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[0] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="App Health :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[2] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="App Status :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[3] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="App Group :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[4] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="App Owner Manager(Penske) :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[6] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="File Total Count :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[7] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Log4j ?  "
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[8] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Secrets? "
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)

                
                item=smart_columns_need_list[10] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Bamboo? "
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[11] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="BitBucket? "
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[12] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Maven? "
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[13] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Module (Not Confirmed) :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[14] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Project Completion Date :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[15] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="CAP Migration (y/n) :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[16] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Original Scope :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[17] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Start Date for Major Blocker :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                item=smart_columns_need_list[18] 
                col=smart_columns_list.index(item) + 1
                font_size=font_format_size[2]
                font_colour=list_of_colour[2]
                val1 = sheet1.cell(row=row, column=col)
                t11 = val1.value
                if(t11==None):
                    t11="No Data Available"
                    font_colour=list_of_colour[1]
                text11="Start Date for Waiting On AO :"
                variable_name.append(text11)
                variable_value.append(t11)
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                give_a_space()
                give_a_space()
                give_a_space()
                
                tspace1 = "         SMARTSHEET COMMENTS        "
                tspaceval = "        "
                variable_name.append(tspace1)
                variable_value.append(tspaceval)
                font_size=font_format_size[0]
                font_colour=list_of_colour[0]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                
                give_a_space()
                give_a_space()
                
                smartsheet_comments(repeat_app_no)
    
    def find_previous_date(exact_date):
        find_date=exact_date[3:5]
        find_mon=exact_date[:2]
        find_year=exact_date[6:8]
        new_date=""
        if(find_date!="01"):
            hj=date_no_list.index(find_date)
            new_date+=exact_date[:3]
            new_date+=date_no_list[hj-1]
            new_date+=exact_date[5:]
            
        elif(find_mon!="01"):
            lj=month_no_list.index(find_mon)
            po=month_last_day_list[lj-1]
            new_date+=po
            new_date+="/"
            po2=month_no_list[lj-1]
            new_date+=po2
            new_date+="/"
            new_date+=exact_date[5:]
            
        else:
            ko1=date_no_list[-1]
            ko2=month_no_list[-1]
            ko3=year_no_list[0]
            new_date+=ko1 + "/" + ko2 + "/" + ko3
            
        return new_date
            
                
    def specific_daywise_update(exact_date):
        prev_date=find_previous_date(exact_date)
        all_matched_date_indexes=[]
        date_count=0
        # Load the Excel file
        workbook = openpyxl.load_workbook('file2.xlsx')

        # Select the desired sheets by index
        second_sheet = workbook.worksheets[1]
        third_sheet = workbook.worksheets[2]

        # Create lists for the first, second, third, and fourth columns
        first_column_values = []
        second_column_values = []
        third_column_values = []
        fourth_column_values = []

        # Iterate over the rows in the second sheet
        row_number_x = 2  # Start from the 2nd row
        for row_xyz in second_sheet.iter_rows(min_row=row_number_x, values_only=True):
            first_value = row_xyz[0]
            second_value = row_xyz[1]
            third_value = row_xyz[2]
            fourth_value = row_xyz[3]

            # Check if the third column has a value and the first column is empty
            if third_value and not first_value:
                # Find the most previous value in the first column
                previous_value = None
                for prev_row in reversed(list(second_sheet.iter_rows(max_row=row_number_x-1, min_col=1, max_col=1, values_only=True))):
                    if prev_row[0]:
                        previous_value = prev_row[0]
                        break
                
                first_column_values.append(previous_value)
            else:
                first_column_values.append(first_value)

            second_column_values.append(second_value)
            third_column_values.append(third_value)
            fourth_column_values.append(fourth_value)

            row_number_x += 1  # Increment the row number

        # Remove None values from the lists
        first_column_values = [value for value in first_column_values if value is not None]
        second_column_values = [value for value in second_column_values if value is not None]
        third_column_values = [value for value in third_column_values if value is not None]
        fourth_column_values2 = [value for value in fourth_column_values if value is not None]
        fourth_column_values1 = [value[:8] if value else None for value in fourth_column_values]
        fourth_column_values = [value for value in fourth_column_values1 if value is not None]

        # Replace 'Row X' values in first_column_values with application names from the first sheet
        first_sheet = workbook.worksheets[0]
        for i, value in enumerate(first_column_values):
            if value.startswith('Row '):
                row_number_x = int(value.split(' ')[1])
                app_name = first_sheet.cell(row=row_number_x+1, column=5).value
                first_column_values[i] = app_name
        
        
        
        commented_application=[]
        comment_message=[]
        comment_person=[]
        comment_date=[]
        ai=0
        for i in fourth_column_values:
            a=str(i)
            if(a[:8]==exact_date or a[:8]==prev_date):# need to check data type
                commented_application.append(first_column_values[ai])
                comment_message.append(second_column_values[ai])
                comment_person.append(third_column_values[ai])
                comment_date.append(fourth_column_values[ai])
            ai+=1
            
        
       
        
        if (len(commented_application)!=0):
            for j in range(len(commented_application)):
                llla=fourth_column_values2[j]
                text_smart=str(comment_person[j])+" added a message for "
                value_smart=str(commented_application[j]) + " application at" + str(llla[8:]) + " on " +str(comment_date[j])+ "\n"
                variable_name.append(text_smart)
                variable_value.append(value_smart)
                
                font_size=font_format_size[4]
                font_colour=list_of_colour[1]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                text_smart="Message:\t "
                value_smart=str(comment_message[j])
                variable_name.append(text_smart)
                variable_value.append(value_smart)
                
                font_size=font_format_size[2]
                font_colour=list_of_colour[0]
                variable_font.append(font_size)
                variable_colour.append(font_colour)
                
                give_a_space()
                give_a_space()
                give_a_space()
                
            
                
        else:
            text_smart="There is no update for any application "
            value_smart=" "
            variable_name.append(text_smart)
            variable_value.append(value_smart)
            
            font_size=font_format_size[2]
            font_colour=list_of_colour[2]
            variable_font.append(font_size)
            variable_colour.append(font_colour)
            give_a_space()
            
        
   
    
    def compare():
        
        def submit_combobox():
            selected_value1 = combo1.get()
            try:
                selected_value2 = combo2.get()
                if(selected_value2=="" or selected_value2==" " or None):
                    selected_value2 = selected_value1
            except EOFError:
                selected_value2 = selected_value1
            smartsheet(selected_value2)
            querytracker(selected_value1)
            for widget in root.winfo_children():
                widget.destroy()
            # create a canvas with a fixed width and height
            canvas = tk.Canvas(root, width=600, height=450)

            # add a vertical scrollbar to the canvas
            v_scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
            canvas.configure(yscrollcommand=v_scrollbar.set)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # add a horizontal scrollbar to the canvas
            h_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.configure(xscrollcommand=h_scrollbar.set)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

            # add some content to the canvas
            content = tk.Frame(canvas)
            len_variable = len(variable_name) 
            
            for i in range(len_variable):
                tk.Label(content, text=f" {variable_name[i]} {variable_value[i]}", justify="left", font=variable_font[i], fg=variable_colour[i]).grid(row=i, column=0, sticky="w")
            
            canvas.create_window((0, 0), window=content, anchor=tk.W)

            # update the canvas to show the content
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            back_button = tk.Button(root, text="HOME", command=clear_window_merge)
            back_button.pack()
            jkl=str(selected_value2)+"_compare_output.txt"
            f = open(jkl, "w")

            for writing in range(len(variable_name)):
                f.write(str(variable_name[writing])+" ")
                f.write(str(variable_value[writing])+"\n")

            # Close the file
            f.close()
            variable_name.clear()
            variable_value.clear()
            variable_font.clear()
            variable_colour.clear()
            
            
        
        def search1(*args):
            query = combo1.get()
            matching_items1 = [item for item in All_apps1 if query.lower() in item.lower()]
            combo1['values'] = matching_items1
        
        
        def search2(*args):
            query = combo2.get()
            matching_items2 = [smart_item for smart_item in All_apps_smart1 if query.lower() in smart_item.lower()]
            combo2['values'] = matching_items2
        
       
        combobox_frame = tk.Frame(root, padx=10, pady=10)
        combobox_frame.pack()
        
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo1 = ttk.Combobox(root, style="Custom.TCombobox", values=All_apps1, state='normal')
        combo1.set("Choose your Query Tracker App")
        combo1.pack(padx=10, pady=10)
        combo1.configure(width=20, height=3)

        combo1.bind("<Key>", search1)

        # Create the second Combobox
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo2 = ttk.Combobox(root, style="Custom.TCombobox", values=All_apps_smart1, state='normal')
        combo2.set("Choose your Smartsheet App")
        combo2.pack(padx=10, pady=10)
        combo2.configure(width=20, height=3)

        combo2.bind("<Key>", search2)

        submit_frame = tk.Frame(root)
        submit_frame.pack(pady=10)

        # Create the Submit button
        submit_button = tk.Button(submit_frame, text="Submit", command=submit_combobox, width=10)
        submit_button.pack()

        # Calculate the window dimensions for centering
        window_width = 400
        window_height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        # Set the window dimensions and position
        root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

        root.mainloop()

    def App_Details():
        
        
        def submit_combobox():
            selected_value1 = combo1.get()
            try:
                selected_value2 = combo2.get()
                if(selected_value2=="" or selected_value2==" " or None):
                    selected_value2 = selected_value1
            except EOFError:
                selected_value2 = selected_value1
            
            check(selected_value1,Apps,app_type)
            log4jv1(selected_value1)
            smartsheet(selected_value2)
            querytracker(selected_value1)
            dsr(selected_value1)
            #smartsheet_comments(selected_value2)
            for widget in root.winfo_children():
                widget.destroy()
            # create a canvas with a fixed width and height
            canvas = tk.Canvas(root, width=600, height=450)

            # add a vertical scrollbar to the canvas
            v_scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
            canvas.configure(yscrollcommand=v_scrollbar.set)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # add a horizontal scrollbar to the canvas
            h_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.configure(xscrollcommand=h_scrollbar.set)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

            # add some content to the canvas
            content = tk.Frame(canvas)
            len_variable = len(variable_name) 
            
            for i in range(len_variable):
                tk.Label(content, text=f" {variable_name[i]} {variable_value[i]}", justify="left", font=variable_font[i], fg=variable_colour[i]).grid(row=i, column=0, sticky="w")
            
            canvas.create_window((0, 0), window=content, anchor=tk.W)

            # update the canvas to show the content
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            back_button = tk.Button(root, text="HOME", command=clear_window_merge)
            back_button.pack()
            jkl=str(selected_value2)+"_app_output.txt"
            f = open(jkl, "w")

            for writing in range(len(variable_name)):
                # Write some text to the file
                f.write(str(variable_name[writing])+"\t")
                f.write(str(variable_value[writing])+"\n")

            # Close the file
            f.close()
            variable_name.clear()
            variable_value.clear()
            variable_font.clear()
            variable_colour.clear()
            
            
            
            
        
        def search1(*args):
            query = combo1.get()
            matching_items1 = [item for item in All_apps1 if query.lower() in item.lower()]
            combo1['values'] = matching_items1
        
        
        def search2(*args):
            query = combo2.get()
            matching_items2 = [smart_item for smart_item in All_apps_smart1 if query.lower() in smart_item.lower()]
            combo2['values'] = matching_items2
        
        combobox_frame = tk.Frame(root, padx=10, pady=10)
        combobox_frame.pack()
        
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo1 = ttk.Combobox(root, style="Custom.TCombobox", values=All_apps1, state='normal')
        combo1.set("Choose your Log4j App")
        combo1.pack(padx=10, pady=10)
        combo1.configure(width=20, height=3)

        combo1.bind("<Key>", search1)

        # Create the second Combobox
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo2 = ttk.Combobox(root, style="Custom.TCombobox", values=All_apps_smart1, state='normal')
        combo2.set("Choose your Smartsheet App")
        combo2.pack(padx=10, pady=10)
        combo2.configure(width=20, height=3)

        combo2.bind("<Key>", search2)

        submit_frame = tk.Frame(root)
        submit_frame.pack(pady=10)

        # Create the Submit button
        submit_button = tk.Button(submit_frame, text="Submit", command=submit_combobox, width=10)
        submit_button.pack()

        # Calculate the window dimensions for centering
        window_width = 400
        window_height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        # Set the window dimensions and position
        root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

        root.mainloop()

        

    def Smartsheet_All():
        
        def submit_combobox():
            selected_value2 = combo2.get()
            log4jv1(selected_value2)
            smartsheet(selected_value2)
            #smartsheet_comments(selected_value2)
            for widget in root.winfo_children():
                widget.destroy()
            # create a canvas with a fixed width and height
            canvas = tk.Canvas(root, width=600, height=450)

            # add a vertical scrollbar to the canvas
            v_scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
            canvas.configure(yscrollcommand=v_scrollbar.set)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # add a horizontal scrollbar to the canvas
            h_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.configure(xscrollcommand=h_scrollbar.set)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

            # add some content to the canvas
            content = tk.Frame(canvas)
            len_variable = len(variable_name) 
            
            for i in range(len_variable):
                tk.Label(content, text=f" {variable_name[i]} {variable_value[i]}", justify="left", font=variable_font[i], fg=variable_colour[i]).grid(row=i, column=0, sticky="w")
            
            canvas.create_window((0, 0), window=content, anchor=tk.W)

            # update the canvas to show the content
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            back_button = tk.Button(root, text="HOME", command=clear_window_merge)
            back_button.pack()
            jkl=str(selected_value2)+"_Smartsheet_app_output.txt"
            f = open(jkl, "w")

            for writing in range(len(variable_name)):
                # Write some text to the file
                f.write(str(variable_name[writing])+"\t")
                f.write(str(variable_value[writing])+"\n")

            # Close the file
            f.close()
            variable_name.clear()
            variable_value.clear()
            variable_font.clear()
            variable_colour.clear()
            
            
        
        def search2(*args):
            query = combo2.get()
            matching_items2 = [smart_item for smart_item in All_apps_smart1 if query.lower() in smart_item.lower()]
            combo2['values'] = matching_items2
        
        
        combobox_frame = tk.Frame(root, padx=10, pady=10)
        combobox_frame.pack()

        # Create the second Combobox
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo2 = ttk.Combobox(root, style="Custom.TCombobox", values=All_apps_smart1, state='normal')
        combo2.set("Choose your Smartsheet App")
        combo2.pack(padx=10, pady=10)
        combo2.configure(width=20, height=3)

        combo2.bind("<Key>", search2)

        submit_frame = tk.Frame(root)
        submit_frame.pack(pady=10)

        # Create the Submit button
        submit_button = tk.Button(submit_frame, text="Submit", command=submit_combobox, width=10)
        submit_button.pack()

        # Calculate the window dimensions for centering
        window_width = 400
        window_height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        # Set the window dimensions and position
        root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

        root.mainloop()
        
    
    def Smartsheet_Day():
        app = QApplication([])
        window = QWidget()
        layout = QVBoxLayout()

        calendar = QCalendarWidget()
        layout.addWidget(calendar)
        

        def get_selected_date():
            date = calendar.selectedDate()
            date_r=date.toString()
            split_date = date_r.split()
            
            selected_date1 = split_date[2]
            dt = dt_list.index(selected_date1)
            selected_date = date_no_list[dt]
            
            selected_month1 = split_date[1]
            mo = mo_list.index(selected_month1)
            selected_month = month_no_list[mo]
            
            selected_year1 = split_date[3]
            ye = year_list.index(selected_year1)
            selected_year = year_no_list[ye]
            
            label = tk.Label(root, text="Smartsheet Daily Update")
            label.pack()
            
            exact_date = str(selected_month)+"/"+str(selected_date)+"/"+str(selected_year)
            
            specific_daywise_update(exact_date)
            
            for widget in root.winfo_children():
                widget.destroy()
            # create a canvas with a fixed width and height
            canvas = tk.Canvas(root, width=600, height=450)

            # add a vertical scrollbar to the canvas
            v_scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
            canvas.configure(yscrollcommand=v_scrollbar.set)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # add a horizontal scrollbar to the canvas
            h_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.configure(xscrollcommand=h_scrollbar.set)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

            # add some content to the canvas
            content = tk.Frame(canvas)
            len_variable = len(variable_name) 
            
            for i in range(len_variable):
                tk.Label(content, text=f" {variable_name[i]} {variable_value[i]}", justify="left", font=variable_font[i], fg=variable_colour[i]).grid(row=i, column=0, sticky="w")
            
            canvas.create_window((0, 0), window=content, anchor=tk.W)

            # update the canvas to show the content
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            back_button = tk.Button(root, text="HOME", command=clear_window_merge)
            back_button.pack()
            jkl=str(exact_date)+"_day_update_output.txt"
            f = open("day_update_output.txt", "w")

            for writing in range(len(variable_name)):
                # Write some text to the file
                f.write(str(variable_name[writing])+"\t")
                f.write(str(variable_value[writing])+"\n")

            # Close the file
            f.close()
            variable_name.clear()
            variable_value.clear()
            variable_font.clear()
            variable_colour.clear()

        calendar.selectionChanged.connect(get_selected_date)

        window.setLayout(layout)
        window.show()

        app.exec_()
        
        
    def JAR_Details():
        
        def submit_combobox():
            selected_value1 = combo1.get()
            jar(selected_value1)
            for widget in root.winfo_children():
                widget.destroy()
            # create a canvas with a fixed width and height
            canvas = tk.Canvas(root, width=600, height=450)

            # add a vertical scrollbar to the canvas
            v_scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
            canvas.configure(yscrollcommand=v_scrollbar.set)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # add a horizontal scrollbar to the canvas
            h_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.configure(xscrollcommand=h_scrollbar.set)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

            # add some content to the canvas
            content = tk.Frame(canvas)
            len_variable = len(variable_name) 
            
            for i in range(len_variable):
                tk.Label(content, text=f" {variable_name[i]} {variable_value[i]}", justify="left", font=variable_font[i], fg=variable_colour[i]).grid(row=i, column=0, sticky="w")
            
            canvas.create_window((0, 0), window=content, anchor=tk.W)

            # update the canvas to show the content
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            back_button = tk.Button(root, text="HOME", command=clear_window_merge)
            back_button.pack()
            jkl=str(selected_value1)+"_jar_output.txt"
            f = open(jkl, "w")

            for writing in range(len(variable_name)):
                # Write some text to the file
                f.write(str(variable_name[writing])+"\t")
                f.write(str(variable_value[writing])+"\n")

            # Close the file
            f.close()
            variable_name.clear()
            variable_value.clear()
            variable_font.clear()
            variable_colour.clear()
            
            
            
            
        
        def search1(*args):
            query = combo1.get()
            matching_items1 = [jar_item for jar_item in All_jars_sheet1 if query.lower() in jar_item.lower()]
            combo1['values'] = matching_items1
        
        
        combobox_frame = tk.Frame(root, padx=10, pady=10)
        combobox_frame.pack()
        
        style = ttk.Style()
        style.configure("Custom.TCombobox", foreground="gray")

        combo1 = ttk.Combobox(root, style="Custom.TCombobox", values=All_jars_sheet1, state='normal')
        combo1.set("Choose your JAR")
        combo1.pack(padx=10, pady=10)
        combo1.configure(width=20, height=3)

        combo1.bind("<Key>", search1)
        

        submit_frame = tk.Frame(root)
        submit_frame.pack(pady=10)

        # Create the Submit button
        submit_button = tk.Button(submit_frame, text="Submit", command=submit_combobox, width=10)
        submit_button.pack()

        # Calculate the window dimensions for centering
        window_width = 400
        window_height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        # Set the window dimensions and position
        root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

        root.mainloop()

    def print_something(val_1):
        label = tk.Label(root, text="You have choose {}".format(val_1))
        label.pack()
        


    def clear_window(val_1):
        for widget in root.winfo_children():
            widget.destroy()
        if(val_1==1):
            compare()
        elif(val_1==2):
            App_Details()
        elif(val_1==3):
            Smartsheet_Day()
        elif(val_1==4):
            JAR_Details()
        else:
            Smartsheet_All()
    
    if(home_val==1):
        clear_window(1)
    elif(home_val==2):
        clear_window(2)
    elif(home_val==3):
        clear_window(3)
    elif(home_val==4):
        clear_window(4)
    else:
        clear_window(5)        


def home():
    selected_option = tk.StringVar()

    option1 = tk.Button(root, text="View Smartsheet & Querytracker\n", command=lambda: common_things(1), bg="light gray", width=25, height=3, font=("Arial", 12))
    option1.grid(row=0, column=0, padx=5, pady=5)

    option2 = tk.Button(root, text="App Details\n", command=lambda: common_things(2), bg="light gray", width=25, height=3, font=("Arial", 12))
    option2.grid(row=1, column=0, padx=5, pady=5)

    option3 = tk.Button(root, text="Daily Update\n", command=lambda: common_things(3), bg="light gray", width=25, height=3, font=("Arial", 12))
    option3.grid(row=2, column=0, padx=5, pady=5)

    option4 = tk.Button(root, text="JAR Details\n", command=lambda: common_things(4), bg="light gray", width=25, height=3, font=("Arial", 12))
    option4.grid(row=3, column=0, padx=5, pady=5)

    option5 = tk.Button(root, text="Smartsheet App Details\n", command=lambda: common_things(5), bg="light gray", width=25, height=3, font=("Arial", 12))
    option5.grid(row=4, column=0, padx=5, pady=5)

    # Configure grid spacing
    root.grid_rowconfigure((0, 1, 2, 3, 4), weight=1)
    root.grid_columnconfigure(0, weight=1)


def clear_window_merge():
    for widget in root.winfo_children():
        widget.destroy()
    home()


root = tk.Tk()
root.title("Excel Extractor Application")

# Change the background color
root.configure(bg="teal")

# Calculate the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the window dimensions for 25% of the screen size
window_width = int(screen_width * 0.25)
window_height = int(screen_height * 0.25)

# Calculate the window position at the center of the screen
window_x = (screen_width - window_width) // 2
window_y = (screen_height - window_height) // 2

# Set the window dimensions and position
root.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

clear_window_merge()
root.mainloop()
